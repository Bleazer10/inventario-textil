from django.http import HttpResponse
from django.contrib.staticfiles import finders

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


def excel_reporte(
    nombre_archivo: str,
    hoja: str,
    titulo: str,
    columnas: list,
    filas: list,
    filtros: list = None,
    resumen: list = None,
    logo_relpath: str = "img/logo-bemore.jpeg",
    nombre_empresa: str = "BEMORE",
    formato_moneda_cols: list = None,  # indices 0-based columnas dinero
):
    wb = Workbook()
    ws = wb.active
    ws.title = hoja[:31]

    last_col = len(columnas)
    last_col_letter = get_column_letter(last_col)

    # ===== Estilos =====
    bold = Font(bold=True)
    company_font = Font(bold=True, size=18)
    title_font = Font(bold=True, size=16)

    header_font = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="111111")
    fill_section = PatternFill("solid", fgColor="F2F2F2")

    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== Logo (más grande) =====
    logo_path = finders.find(logo_relpath)
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.height = 90
            img.width = 90
            ws.add_image(img, "A1")
        except:
            pass

    # ===== Encabezado SIMPLE (como antes, no centrado) =====
    # DESPUÉS (mejor, no se monta con el logo)
    ws["B1"] = nombre_empresa
    ws["B1"].font = company_font
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["B3"] = titulo
    ws["B3"].font = title_font
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 22

    fila_actual = 5

    # ===== Filtros (solo si hay algo) =====
    filtros_visibles = []
    if filtros:
        for f in filtros:
            if isinstance(f, str):
                texto = f.strip()
                if texto not in ("", "—", "-", "None", "null"):
                    filtros_visibles.append([texto])
            elif isinstance(f, (list, tuple)):
                if len(f) == 1:
                    texto = str(f[0]).strip()
                    if texto not in ("", "—", "-", "None", "null"):
                        filtros_visibles.append([texto])
                elif len(f) >= 2:
                    k = str(f[0]).strip()
                    v = str(f[1]).strip()
                    if v not in ("", "—", "-", "None", "null"):
                        filtros_visibles.append([k, v])

    if filtros_visibles:
        for row in filtros_visibles:
            if len(row) == 1:
                texto = row[0]
                ws[f"A{fila_actual}"] = texto
                ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=len(columnas))
                ws[f"A{fila_actual}"].fill = fill_section
                ws[f"A{fila_actual}"].border = border
                ws[f"A{fila_actual}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                k, v = row
                ws[f"A{fila_actual}"] = k
                ws[f"B{fila_actual}"] = v
                ws[f"A{fila_actual}"].fill = fill_section
                ws[f"A{fila_actual}"].border = border
                ws[f"B{fila_actual}"].border = border
            fila_actual += 1

        fila_actual += 1

    # ===== Resumen (como antes, no centrado) =====
    if resumen:
        ws[f"A{fila_actual}"] = "Resumen"
        ws[f"A{fila_actual}"].font = bold
        fila_actual += 1

        for k, v in resumen:
            ws[f"A{fila_actual}"] = k
            ws[f"B{fila_actual}"] = v
            ws[f"A{fila_actual}"].fill = fill_section
            ws[f"A{fila_actual}"].border = border
            ws[f"B{fila_actual}"].border = border
            fila_actual += 1

        fila_actual += 2

    # ===== Tabla principal =====
    fila_header = fila_actual

    # Encabezados
    for col_idx, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=fila_header, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[fila_header].height = 20

    # Datos
    for r in filas:
        fila_actual += 1
        for col_idx, val in enumerate(r, start=1):
            c = ws.cell(row=fila_actual, column=col_idx, value=val)
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center")

    fila_fin = fila_actual

    # Congelar encabezado (esto sí es útil y no pone flechas)
    ws.freeze_panes = ws[f"A{fila_header+1}"]

    # ✅ QUITAR AUTO-FILTER (esto elimina la flecha gris)
    # ws.auto_filter.ref = f"A{fila_header}:{last_col_letter}{fila_fin}"  <-- NO

    # Formato moneda (alineado derecha)
    if formato_moneda_cols:
        for col0 in formato_moneda_cols:
            col = col0 + 1
            for row in range(fila_header + 1, fila_fin + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ===== Anchos fijos buenos (sin locuras) =====
    # Ajusta si quieres, pero estos ya quedan bien para ventas.
    anchos = {
        1: 20,   # #Venta
        2: 14,  # Fecha
        3: 24,  # Cliente
        4: 12,  # Tipo pago
        5: 12,  # Total
        6: 12,  # Pagado
        7: 12,  # Saldo
        8: 14,  # Estado
    }
    for col in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = anchos.get(col, 16)

    # ===== Respuesta HTTP =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response